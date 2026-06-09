"""
Carga histórica TRM desde 2020-01-01 hasta hoy.
Expande rangos vigenciadesde/vigenciahasta para cubrir TODAS las fechas del calendario.
Ejecutar UNA SOLA VEZ para inicializar el Excel.
"""

import requests
import json
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "TRM_USD_COP diario.xlsx"
JSON_PATH  = BASE_DIR / "trm_data.json"
LOG_PATH   = BASE_DIR / ".tmp" / "error_log.txt"
API_URL    = "https://www.datos.gov.co/resource/ceyp-9c7c.json"

def log_error(msg):
    LOG_PATH.parent.mkdir(exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().isoformat()}] {msg}\n")

def fetch_historical():
    params = {
        "$where": "vigenciadesde >= '2020-01-01T00:00:00.000'",
        "$order": "vigenciadesde ASC",
        "$limit": 10000
    }
    try:
        r = requests.get(API_URL, params=params, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log_error(f"Error fetching historical data: {e}")
        return []

def expand_records(raw):
    """Expande registros usando el más reciente como prioridad.
    Fechas sin dato hasta hoy se rellenan con el último TRM conocido."""
    expanded = {}
    hoy = date.today()

    # Paso 1: procesar de más reciente a más antiguo — el registro más nuevo manda
    raw_sorted = sorted(raw, key=lambda r: r["vigenciadesde"], reverse=True)
    for rec in raw_sorted:
        trm   = round(float(rec["valor"]), 2)
        desde = date.fromisoformat(rec["vigenciadesde"][:10])
        hasta = date.fromisoformat(rec["vigenciahasta"][:10])
        d = desde
        while d <= hasta and d <= hoy:
            k = d.isoformat()
            if k not in expanded:
                expanded[k] = trm
            d += timedelta(days=1)

    # Paso 2: rellenar huecos hasta hoy con el último TRM conocido
    if expanded:
        ultima_fecha = date.fromisoformat(max(expanded.keys()))
        ultimo_trm   = expanded[ultima_fecha.isoformat()]
        d = ultima_fecha + timedelta(days=1)
        while d <= hoy:
            expanded[d.isoformat()] = ultimo_trm
            d += timedelta(days=1)

    return sorted(expanded.items())  # [(fecha_str, trm), ...]

def create_excel_with_style(wb):
    ws = wb.active
    ws.title = "TRM Diario"
    header_fill  = PatternFill("solid", fgColor="2E7D32")
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin   = Side(border_style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(["Fecha", "TRM (COP/USD)"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = border
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.row_dimensions[1].height = 25
    return ws

def write_excel(records):
    # Siempre recrea el archivo limpio para garantizar orden cronológico
    wb = openpyxl.Workbook()
    ws = create_excel_with_style(wb)

    thin        = Side(border_style="thin", color="BDBDBD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even   = PatternFill("solid", fgColor="E8F5E9")

    for i, (fecha, trm) in enumerate(records, start=2):
        cf = ws.cell(row=i, column=1, value=fecha)
        ct = ws.cell(row=i, column=2, value=trm)
        for cell in [cf, ct]:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = fill_even
        ct.number_format = '#,##0.00'

    wb.save(EXCEL_PATH)
    print(f"Excel guardado: {EXCEL_PATH} ({len(records)} registros)")

def write_json(records):
    data = [{"fecha": f, "trm": t} for f, t in records]
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"JSON guardado: {JSON_PATH} ({len(data)} registros)")

def main():
    print("Iniciando carga histórica TRM desde 2020-01-01...")
    raw = fetch_historical()
    if not raw:
        print("No se obtuvieron datos. Revisa .tmp/error_log.txt")
        return
    print(f"Registros brutos de la API: {len(raw)}")
    records = expand_records(raw)
    print(f"Registros expandidos (todas las fechas): {len(records)}")
    write_excel(records)
    write_json(records)
    print(f"Primer registro: {records[0][0]}  |  Último: {records[-1][0]}")
    print("Carga histórica completada.")

if __name__ == "__main__":
    main()

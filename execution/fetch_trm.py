"""
Actualización diaria TRM. Se ejecuta cada día a las 11:30 AM (COT).
Expande rangos para cubrir todas las fechas del calendario.
"""

import requests
import json
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side

BASE_DIR   = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "TRM_USD_COP diario.xlsx"
JSON_PATH  = BASE_DIR / "trm_data.json"
LOG_PATH   = BASE_DIR / ".tmp" / "error_log.txt"
API_URL    = "https://www.datos.gov.co/resource/ceyp-9c7c.json"

def log_error(msg):
    LOG_PATH.parent.mkdir(exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().isoformat()}] {msg}\n")

def fetch_latest():
    params = {"$limit": 5, "$order": "vigenciadesde DESC"}
    try:
        r = requests.get(API_URL, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        if not data:
            raise ValueError("API returned empty list")
        return data
    except Exception as e:
        log_error(f"fetch_latest error: {e}")
        return []

def expand_records(raw):
    ayer = date.today() - timedelta(days=1)  # solo hasta ayer
    expanded = {}
    raw_sorted = sorted(raw, key=lambda r: r["vigenciadesde"], reverse=True)
    for rec in raw_sorted:
        trm   = round(float(rec["valor"]), 2)
        desde = date.fromisoformat(rec["vigenciadesde"][:10])
        hasta = date.fromisoformat(rec["vigenciahasta"][:10])
        d = desde
        while d <= hasta and d <= ayer:
            k = d.isoformat()
            if k not in expanded:
                expanded[k] = trm
            d += timedelta(days=1)
    # Rellenar huecos hasta ayer
    if expanded:
        ultima = date.fromisoformat(max(expanded.keys()))
        ultimo_trm = expanded[ultima.isoformat()]
        d = ultima + timedelta(days=1)
        while d <= ayer:
            expanded[d.isoformat()] = ultimo_trm
            d += timedelta(days=1)
    return expanded  # {fecha_str: trm}

def append_to_excel(new_records):
    if not EXCEL_PATH.exists():
        log_error(f"Excel no encontrado. Ejecuta historical_load.py primero.")
        return 0

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(str(row[0]))

    thin      = Side(border_style="thin", color="BDBDBD")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill("solid", fgColor="E8F5E9")

    added = 0
    for fecha in sorted(new_records):
        if fecha in existing:
            continue
        trm      = new_records[fecha]
        next_row = ws.max_row + 1
        cf = ws.cell(row=next_row, column=1, value=fecha)
        ct = ws.cell(row=next_row, column=2, value=trm)
        for cell in [cf, ct]:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if next_row % 2 == 0:
                cell.fill = fill_even
        ct.number_format = '#,##0.00'
        existing.add(fecha)
        added += 1

    if added:
        wb.save(EXCEL_PATH)
    return added

def update_json(new_records):
    data = []
    if JSON_PATH.exists():
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

    existing = {r["fecha"] for r in data}
    for fecha, trm in new_records.items():
        if fecha not in existing:
            data.append({"fecha": fecha, "trm": trm})

    data.sort(key=lambda x: x["fecha"])
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    return len(new_records) - len(existing.intersection(new_records))

def push_to_github():
    import subprocess
    repo_dir = BASE_DIR
    try:
        subprocess.run(["git", "-C", str(repo_dir), "add", "trm_data.json"], check=True)
        subprocess.run(["git", "-C", str(repo_dir), "commit", "-m", f"TRM actualizado {date.today().isoformat()}"], check=True)
        subprocess.run(["git", "-C", str(repo_dir), "push"], check=True)
        print("Dashboard publicado en GitHub Pages.")
    except subprocess.CalledProcessError as e:
        log_error(f"git push error: {e}")
        print("Advertencia: no se pudo subir a GitHub. Ver error_log.txt")

def main():
    raw = fetch_latest()
    if not raw:
        print("No se pudo obtener TRM. Ver error_log.txt")
        return

    new_records = expand_records(raw)
    added_excel = append_to_excel(new_records)
    update_json(new_records)

    latest = sorted(new_records.items())[-1]
    print(f"TRM {latest[0]}: ${latest[1]:,.2f} COP")
    print(f"Registros nuevos agregados: {added_excel}")

    if added_excel > 0:
        push_to_github()

if __name__ == "__main__":
    main()
